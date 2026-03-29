# src/jaegerbot/analysis/show_results.py (Final Version 9 - Fix für Modus 2)
import os
import sys
import json
import pandas as pd
import numpy as np # Import für np.nan
from datetime import date, datetime
import logging
import argparse

logging.getLogger('tensorflow').setLevel(logging.ERROR)
logging.getLogger('absl').setLevel(logging.ERROR)

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
sys.path.append(os.path.join(PROJECT_ROOT, 'src'))

from jaegerbot.analysis.backtester import load_data
from jaegerbot.utils.ann_model import load_model_and_scaler
from jaegerbot.analysis.portfolio_simulator import run_portfolio_simulation
from jaegerbot.analysis.portfolio_optimizer import run_portfolio_optimizer
from jaegerbot.utils.telegram import send_document

GREEN = '\033[0;32m'
YELLOW = '\033[1;33m'
NC = '\033[0m'


def _get_telegram_cfg():
    try:
        with open(os.path.join(PROJECT_ROOT, 'secret.json'), 'r') as f:
            s = json.load(f)
        tg = s.get('telegram', {})
        return tg.get('bot_token', ''), tg.get('chat_id', '')
    except Exception:
        return '', ''


def _generate_trades_excel(final_sim, capital):
    """Erstellt jaegerbot_trades.xlsx mit allen Portfolio-Trades — analog zu vbot/fibot."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"  {YELLOW}openpyxl nicht installiert — Excel übersprungen. (pip install openpyxl){NC}")
        return

    trade_history = final_sim.get('trade_history', [])
    if not trade_history:
        print(f"  {YELLOW}Keine Trades — Excel übersprungen.{NC}")
        return

    equity = capital
    rows   = []
    for i, t in enumerate(trade_history):
        pnl      = float(t['pnl'])
        equity  += pnl
        symbol   = t.get('symbol', '')
        tf       = t.get('timeframe', '')
        strat    = f"{symbol.split('/')[0]}/{tf}" if symbol else tf
        dir_     = t.get('direction', '').upper()
        entry    = round(float(t.get('entry', 0)), 6)
        exit_p   = round(float(t.get('exit',  0)), 6)
        ergebnis = 'TP erreicht' if pnl > 0 else 'SL erreicht'
        rows.append({
            'Nr':           i + 1,
            'Datum':        str(t.get('entry_time', t.get('ts', '')))[:16].replace('T', ' '),
            'Strategie':    strat,
            'Richtung':     dir_,
            'Hebel':        t.get('leverage', '—'),
            'Einsatz (USDT)': round(float(t.get('margin_used', 0)), 2),
            'Entry':        entry,
            'Exit':         exit_p,
            'Ergebnis':     ergebnis,
            'PnL (USDT)':   round(pnl,    4),
            'Kapital':      round(equity, 4),
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trades'

    header_fill = PatternFill('solid', fgColor='1E3A5F')
    win_fill    = PatternFill('solid', fgColor='D6F4DC')
    loss_fill   = PatternFill('solid', fgColor='FAD7D7')
    alt_fill    = PatternFill('solid', fgColor='F2F2F2')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
    )
    col_widths = {
        'Nr': 5, 'Datum': 18, 'Strategie': 22, 'Richtung': 10,
        'Hebel': 8, 'Einsatz (USDT)': 16,
        'Entry': 14, 'Exit': 14, 'Ergebnis': 14, 'PnL (USDT)': 14, 'Kapital': 16,
    }

    headers = list(rows[0].keys())
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(h, 14)
    ws.row_dimensions[1].height = 22

    for r_idx, row in enumerate(rows, 2):
        if row['Ergebnis'] == 'TP erreicht':
            fill = win_fill
        elif r_idx % 2 == 0:
            fill = loss_fill
        else:
            fill = alt_fill
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=col, value=row[key])
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('Entry', 'Exit', 'PnL (USDT)', 'Kapital'):
                cell.number_format = '#,##0.0000'
            elif key == 'Einsatz (USDT)':
                cell.number_format = '#,##0.00'
        ws.row_dimensions[r_idx].height = 18

    total     = len(rows)
    wins      = sum(1 for r in rows if r['Ergebnis'] == 'TP erreicht')
    sr        = total + 3
    pnl_total = rows[-1]['Kapital'] - capital if rows else 0.0
    pnl_pct   = pnl_total / capital * 100 if capital else 0.0
    ws.cell(row=sr, column=1, value='Zusammenfassung').font = Font(bold=True, size=11)
    for label, value in [
        ('Trades gesamt', total),
        ('Win-Rate',      f"{wins / total * 100:.1f}%" if total else '—'),
        ('PnL',           f"{pnl_pct:+.1f}%"),
        ('Endkapital',    f"{rows[-1]['Kapital']:.2f} USDT" if rows else '—'),
    ]:
        ws.cell(row=sr, column=1, value=label).font = Font(bold=True)
        ws.cell(row=sr, column=2, value=value)
        sr += 1

    out_dir  = os.path.join(PROJECT_ROOT, 'artifacts', 'charts')
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, 'jaegerbot_trades.xlsx')
    wb.save(out_file)
    print(f"  {GREEN}Excel gespeichert: jaegerbot_trades.xlsx{NC}")

    bot_token, chat_id = _get_telegram_cfg()
    if bot_token and chat_id:
        caption = (f"JaegerBot Trades — {total} Trades | "
                   f"WR: {wins / total * 100:.1f}% | PnL: {pnl_pct:+.1f}%" if total else "JaegerBot Trades")
        send_document(bot_token, chat_id, out_file, caption=caption)
        print(f"  {GREEN}Via Telegram gesendet.{NC}")
    else:
        print(f"  {YELLOW}Telegram nicht konfiguriert — nur lokal gespeichert.{NC}")


def _generate_portfolio_chart(final_sim, portfolio_files, capital, start_date, end_date):
    """Erstellt jaegerbot_portfolio_equity.html — analog zu fibot_portfolio_equity.html."""
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError:
        print(f"  {YELLOW}plotly nicht installiert — Chart übersprungen. (pip install plotly){NC}")
        return

    eq_df         = final_sim.get('equity_curve')
    trade_history = final_sim.get('trade_history', [])
    if eq_df is None or (hasattr(eq_df, 'empty') and eq_df.empty):
        print(f"  {YELLOW}Keine Equity-Daten — Chart übersprungen.{NC}")
        return

    eq_times = eq_df['timestamp'].astype(str).tolist()
    eq_vals  = eq_df['equity'].tolist()

    # Trade-Marker: Win ● cyan / Loss ✗ rot
    win_x, win_y   = [], []
    loss_x, loss_y = [], []
    for t in trade_history:
        ts  = str(t.get('ts', ''))
        row = eq_df[eq_df['timestamp'] <= pd.to_datetime(t['ts'])]
        eq_at = float(row['equity'].iloc[-1]) if not row.empty else capital
        if float(t['pnl']) > 0:
            win_x.append(ts);  win_y.append(eq_at)
        else:
            loss_x.append(ts); loss_y.append(eq_at)

    # Strategie-Labels für Titel
    pairs = []
    for fname in portfolio_files:
        name  = fname.replace('config_', '').replace('.json', '')
        parts = name.split('_')
        tf    = parts[-1] if parts else ''
        sym   = parts[0][:3].upper() if parts else ''
        pairs.append(f"{sym}/{tf}")
    pairs_str = ', '.join(pairs)

    n_strats = len(portfolio_files)
    pnl_pct  = final_sim.get('total_pnl_pct', 0)
    sign     = '+' if pnl_pct >= 0 else ''
    title = (
        f"JaegerBot Portfolio — {n_strats} Strategie(n) ({pairs_str}) | "
        f"Zeitraum: {start_date} \u2192 {end_date} | "
        f"Trades: {final_sim.get('trade_count', 0)} | WR: {final_sim.get('win_rate', 0):.1f}% | "
        f"PnL: {sign}{pnl_pct:.1f}% | "
        f"Endkapital: {final_sim.get('end_capital', capital):.2f} USDT | "
        f"MaxDD: {final_sim.get('max_drawdown_pct', 0):.1f}%"
    )

    fig = make_subplots(specs=[[{"secondary_y": False}]])

    fig.add_hline(
        y=capital,
        line=dict(color='rgba(100,100,100,0.35)', width=1, dash='dash'),
        annotation_text=f'Start {capital:.0f} USDT',
        annotation_position='top left',
    )

    fig.add_trace(go.Scatter(
        x=eq_times, y=eq_vals,
        mode='lines', name='Portfolio Equity',
        line=dict(color='#2563eb', width=2.5),
        hovertemplate='Portfolio: %{y:.2f} USDT<extra></extra>',
    ))

    if win_x:
        fig.add_trace(go.Scatter(
            x=win_x, y=win_y, mode='markers',
            marker=dict(color='#22d3ee', symbol='circle', size=8,
                        line=dict(width=1, color='#0e7490')),
            name='TP \u2713',
        ))

    if loss_x:
        fig.add_trace(go.Scatter(
            x=loss_x, y=loss_y, mode='markers',
            marker=dict(color='#ef4444', symbol='x', size=8,
                        line=dict(width=2, color='#7f1d1d')),
            name='SL \u2717',
        ))

    fig.update_layout(
        title=dict(text=title, font=dict(size=12), x=0.5, xanchor='center'),
        height=600,
        hovermode='x unified',
        template='plotly_dark',
        dragmode='zoom',
        xaxis=dict(rangeslider=dict(visible=True), fixedrange=False),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
        margin=dict(l=60, r=60, t=80, b=40),
        yaxis=dict(title='Equity (USDT)', fixedrange=False),
    )

    out_dir  = os.path.join(PROJECT_ROOT, 'artifacts', 'charts')
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, 'jaegerbot_portfolio_equity.html')
    fig.write_html(out_file)
    print(f"  {GREEN}Chart gespeichert: jaegerbot_portfolio_equity.html{NC}")

    bot_token, chat_id = _get_telegram_cfg()
    if bot_token and chat_id:
        caption = (
            f"JaegerBot Portfolio-Equity\n"
            f"{start_date} \u2192 {end_date} | {n_strats} Strategie(n) | "
            f"PnL: {sign}{pnl_pct:.1f}% | Equity: {final_sim.get('end_capital', capital):.2f} USDT | "
            f"MaxDD: {final_sim.get('max_drawdown_pct', 0):.1f}%"
        )
        send_document(bot_token, chat_id, out_file, caption=caption)
        print(f"  {GREEN}Chart via Telegram gesendet.{NC}")
    else:
        print(f"  {YELLOW}Telegram nicht konfiguriert — Chart nur lokal gespeichert.{NC}")

# --- Helper-Funktion für die Einzelanalyse (Modus 1) ---
def run_single_analysis_via_simulator(start_date, end_date, start_capital):
    print("--- JaegerBot Ergebnis-Analyse (Einzel-Modus via Simulator) ---")
    configs_dir = os.path.join(PROJECT_ROOT, 'src', 'jaegerbot', 'strategy', 'configs')
    models_dir = os.path.join(PROJECT_ROOT, 'artifacts', 'models')
    all_results = []
    
    config_files = sorted([f for f in os.listdir(configs_dir) if f.startswith('config_') and f.endswith('.json')])

    if not config_files:
        print("\nKeine gültigen Konfigurationen zum Analysieren gefunden."); return

    for filename in config_files:
        config_path = os.path.join(configs_dir, filename)
        if not os.path.exists(config_path): continue

        with open(config_path, 'r') as f: config = json.load(f)

        symbol, timeframe = config['market']['symbol'], config['market']['timeframe']
        strategy_name = f"{symbol} ({timeframe})"

        print(f"\nAnalysiere Ergebnisse für: {strategy_name}...")

        safe_filename = f"{symbol.replace('/', '').replace(':', '')}_{timeframe}"
        model_paths = {
            'model': os.path.join(models_dir, f'ann_predictor_{safe_filename}.h5'),
            'scaler': os.path.join(models_dir, f'ann_scaler_{safe_filename}.joblib')
        }

        if not os.path.exists(model_paths['model']):
            print(f"--> WARNUNG: Modell nicht gefunden. Überspringe."); continue

        data = load_data(symbol, timeframe, start_date, end_date)
        if data.empty:
            print(f"--> WARNUNG: Konnte keine Daten laden. Überspringe."); continue

        # Nur eine Strategie in das Dict laden
        strategies_data = {
            symbol: {
                'symbol': symbol, 
                'timeframe': timeframe, 
                'data': data, 
                'model': load_model_and_scaler(model_paths['model'], model_paths['scaler'])[0], 
                'scaler': load_model_and_scaler(model_paths['model'], model_paths['scaler'])[1], 
                'params': {**config.get('strategy', {}), **config.get('risk', {})}
            }
        }
        
        # Führe den Portfolio-Simulator nur für diese eine Strategie aus
        result = run_portfolio_simulation(start_capital, strategies_data, start_date, end_date)
        
        if result['trade_count'] == 0:
            pnl_value = 0.0
            end_capital_value = start_capital
        else:
            pnl_value = result['total_pnl_pct']
            end_capital_value = result['end_capital']

        all_results.append({
            "Strategie": strategy_name,
            "Trades": result['trade_count'],
            "Win Rate %": result.get('win_rate', 0.0),
            "PnL %": pnl_value,
            "Max DD %": result['max_drawdown_pct'],
            "Endkapital": end_capital_value
        })

    if not all_results:
        print("\nKeine gültigen Konfigurationen mehr übrig."); return

    results_df = pd.DataFrame(all_results)
    results_df = results_df.sort_values(by="PnL %", ascending=False)

    display_columns = ["Strategie", "Trades", "Win Rate %", "PnL %", "Max DD %", "Endkapital"]

    pd.set_option('display.width', 1000); pd.set_option('display.max_columns', None)
    print("\n\n==================================================================================");
    print(f"                 Zusammenfassung (Startkapital: {start_capital} USDT)");
    print("==================================================================================")
    pd.set_option('display.float_format', '{:.2f}'.format);
    print(results_df.fillna('-').to_string(index=False, columns=display_columns));
    print("==================================================================================")
# --- ENDE Helper-Funktion ---


def run_shared_mode(is_auto: bool, start_date, end_date, start_capital, max_drawdown=100.0):
    mode_name = "Automatische Portfolio-Optimierung" if is_auto else "Manuelle Portfolio-Simulation"
    print(f"--- JaegerBot {mode_name} ---")
    configs_dir = os.path.join(PROJECT_ROOT, 'src', 'jaegerbot', 'strategy', 'configs')
    models_dir = os.path.join(PROJECT_ROOT, 'artifacts', 'models')
    available_strategies = []
    
    # Lade Liste der verfügbaren Strategien
    if os.path.isdir(configs_dir):
        for filename in sorted(os.listdir(configs_dir)):
            if filename.startswith('config_') and filename.endswith('.json'):
                base_name = filename.replace('config_', '').replace('.json', '').replace('_macd', '')
                try:
                    parts = base_name.split('_'); timeframe = parts[-1]; symbol_part = "_".join(parts[:-1])
                    model_name = f"ann_predictor_{symbol_part}_{timeframe}.h5"
                    if os.path.exists(os.path.join(models_dir, model_name)):
                        available_strategies.append(filename)
                except IndexError: continue
    
    if not available_strategies:
        print("Keine optimierten Strategien gefunden."); return

    selected_files = []
    if not is_auto:
        # --- Modus 2: Manuelle Auswahl ---
        print("\nVerfügbare Strategien:")
        for i, name in enumerate(available_strategies): print(f"  {i+1}) {name}")
        selection = input("\nWelche Strategien sollen simuliert werden? (Zahlen mit Komma, z.B. 1,3,4 oder 'alle'): ")
        try:
            if selection.lower() == 'alle': selected_files = available_strategies
            else: selected_files = [available_strategies[int(i.strip()) - 1] for i in selection.split(',')]
        except (ValueError, IndexError): print("Ungültige Auswahl. Breche ab."); return
    else: 
        # --- Modus 3: Automatische Auswahl ---
        selected_files = available_strategies

    strategies_data = {}
    print("\nLade Daten und Modelle für gewählte Strategien...")
    for filename in selected_files:
        with open(os.path.join(configs_dir, filename), 'r') as f: config = json.load(f)
        symbol, timeframe = config['market']['symbol'], config['market']['timeframe']
        safe_filename = f"{symbol.replace('/', '').replace(':', '')}_{timeframe}"
        
        # Versuche, Modell und Scaler zu laden
        model_paths = {'model': os.path.join(models_dir, f'ann_predictor_{safe_filename}.h5'), 'scaler': os.path.join(models_dir, f'ann_scaler_{safe_filename}.joblib')}
        try:
            model, scaler = load_model_and_scaler(model_paths['model'], model_paths['scaler'])
        except Exception as e:
            print(f"WARNUNG: Konnte Modell/Scaler für {filename} nicht laden. Fehler: {e}. Wird ignoriert.")
            continue
            
        data = load_data(symbol, timeframe, start_date, end_date)
        if model and scaler and not data.empty:
            strategies_data[symbol] = {'symbol': symbol, 'timeframe': timeframe, 'data': data, 'model': model, 'scaler': scaler, 'params': {**config.get('strategy', {}), **config.get('risk', {})}}
        else:
            print(f"WARNUNG: Konnte Daten/Modell für {filename} nicht laden. Wird ignoriert.")
    
    if not strategies_data:
        print("Konnte für keine der gewählten Strategien Daten laden. Breche ab."); return

    # Variablen für das Reporting (werden in beiden Modi gefüllt)
    equity_df = pd.DataFrame()
    csv_path = ""
    caption = ""
    trade_count = 0
    
    try:
        d1 = datetime.strptime(start_date, "%Y-%m-%d")
        d2 = datetime.strptime(end_date, "%Y-%m-%d")
        total_days = (d2 - d1).days
        if total_days <= 0: total_days = 1
    except Exception:
        total_days = 0

    if is_auto:
        # --- Modus 3: Automatische Optimierung ---
        print(f"\nINFO: Starte Optimierung mit maximal {max_drawdown:.2f}% Drawdown-Beschränkung.")

        # Jede Config bekommt ihre eigenen Daten (Symbol+Timeframe-spezifisch),
        # damit BTC/15m und BTC/6h nicht gegenseitig überschrieben werden.
        strategies_data_for_optimizer = {}
        for filename in selected_files:
            try:
                with open(os.path.join(configs_dir, filename), 'r') as f: config = json.load(f)
                symbol    = config['market']['symbol']
                timeframe = config['market']['timeframe']
                safe_fn   = f"{symbol.replace('/', '').replace(':', '')}_{timeframe}"
                m_path    = os.path.join(models_dir, f'ann_predictor_{safe_fn}.h5')
                s_path    = os.path.join(models_dir, f'ann_scaler_{safe_fn}.joblib')
                model, scaler = load_model_and_scaler(m_path, s_path)
                data = load_data(symbol, timeframe, start_date, end_date)
                if model and scaler and not data.empty:
                    strategies_data_for_optimizer[filename] = {
                        'symbol': symbol, 'timeframe': timeframe, 'data': data,
                        'model': model, 'scaler': scaler,
                        'params': {**config.get('strategy', {}), **config.get('risk', {})}
                    }
                else:
                    print(f"WARNUNG: Konnte Daten/Modell für {filename} nicht laden. Wird ignoriert.")
            except Exception as e:
                print(f"WARNUNG: {filename} — {e}. Wird ignoriert.")

        results = run_portfolio_optimizer(start_capital, strategies_data_for_optimizer, start_date, end_date, max_drawdown)

        if results and 'final_result' in results:
            final_report = results['final_result']

            trade_count = final_report.get('trade_count', 0)
            
            # Reporting-Strings
            days_per_trade_str = ""
            if trade_count > 0 and total_days > 0:
                days_per_trade = total_days / trade_count
                days_per_trade_str = f" (entspricht 1 Trade alle {days_per_trade:.1f} Tage)"

            print("\n======================================================="); print("     Ergebnis der automatischen Portfolio-Optimierung"); print("=======================================================")
            print(f"Zeitraum: {start_date} bis {end_date} ({total_days} Tage)\nStartkapital: {start_capital:.2f} USDT")
            print(f"Maximal erlaubter DD: {max_drawdown:.2f}%")
            print("\nOptimales Portfolio gefunden (" + str(len(results['optimal_portfolio'])) + " Strategien):")
            for strat_filename in results['optimal_portfolio']: print(f"  - {strat_filename}")
            print("\n--- Simulierte Performance dieses optimalen Portfolios ---")
            print(f"Endkapital:       {final_report['end_capital']:.2f} USDT"); print(f"Gesamt PnL:       {final_report['end_capital'] - start_capital:+.2f} USDT ({final_report['total_pnl_pct']:.2f}%)")
            print(f"Anzahl Trades:    {trade_count}{days_per_trade_str}")
            print(f"Portfolio Max DD:   {final_report['max_drawdown_pct']:.2f}%")
            print(f"Liquidiert:       {'JA, am ' + final_report['liquidation_date'].strftime('%Y-%m-%d') if final_report['liquidation_date'] else 'NEIN'}")

            # Speichere optimale Strategien für das Bash-Script
            optimal_configs_file = os.path.join(PROJECT_ROOT, '.optimal_configs.tmp')
            with open(optimal_configs_file, 'w') as f:
                f.write('\n'.join(results['optimal_portfolio']))
            
            csv_path = os.path.join(PROJECT_ROOT, 'optimal_portfolio_equity.csv')
            caption = f"Automatischer Portfolio-Optimierungsbericht\nMax. erlaubter DD: {max_drawdown:.2f}%\nTrades: {trade_count}{days_per_trade_str}\nEndkapital: {final_report['end_capital']:.2f} USDT"
            equity_df = final_report.get('equity_curve')
        else:
            print("\n======================================================="); print("     Ergebnis der automatischen Portfolio-Optimierung"); print("=======================================================")
            print(f"❌ Es konnte kein Portfolio gefunden werden, das die Drawdown-Beschränkung von {max_drawdown:.2f}% erfüllt.")
            
    else:
        # --- Modus 2: Manuelle Portfolio-Simulation (KORRIGIERT) ---
        print("\n--- Starte Manuelle Portfolio-Simulation... ---")
        
        final_report = run_portfolio_simulation(start_capital, strategies_data, start_date, end_date)
        
        trade_count = final_report.get('trade_count', 0)
        
        days_per_trade_str = ""
        if trade_count > 0 and total_days > 0:
            days_per_trade = total_days / trade_count
            days_per_trade_str = f" (entspricht 1 Trade alle {days_per_trade:.1f} Tage)"

        print("\n======================================================="); 
        print("     Ergebnis der Manuellen Portfolio-Simulation"); 
        print("=======================================================")
        print(f"Zeitraum: {start_date} bis {end_date} ({total_days} Tage)\nStartkapital: {start_capital:.2f} USDT")
        print(f"Simulierte Strategien: {len(strategies_data)}")
        for key in strategies_data.keys():
            print(f"  - {key}")
            
        print("\n--- Simulierte Performance dieses Portfolios ---")
        print(f"Endkapital:       {final_report['end_capital']:.2f} USDT"); 
        print(f"Gesamt PnL:       {final_report['end_capital'] - start_capital:+.2f} USDT ({final_report['total_pnl_pct']:.2f}%)")
        print(f"Anzahl Trades:    {trade_count}{days_per_trade_str}")
        print(f"Portfolio Max DD:   {final_report['max_drawdown_pct']:.2f}%")
        print(f"Liquidiert:       {'JA, am ' + final_report['liquidation_date'].strftime('%Y-%m-%d') if final_report['liquidation_date'] else 'NEIN'}")
        
        csv_path = os.path.join(PROJECT_ROOT, 'manual_portfolio_equity.csv')
        caption = f"Manueller Portfolio-Bericht\nStrategien: {len(strategies_data)}\nTrades: {trade_count}{days_per_trade_str}\nEndkapital: {final_report['end_capital']:.2f} USDT"
        equity_df = final_report.get('equity_curve')
        # --- ENDE: Manuelle Portfolio-Simulation ---


    # --- Export-Logik (wird für Modus 2 und 3 verwendet) ---
    if equity_df is not None and not equity_df.empty:
        print("\n--- Export ---")
        if not is_auto:
            equity_df[['timestamp', 'equity', 'drawdown_pct']].to_csv(csv_path, index=False)
            print(f"✔ Bericht wurde erfolgreich an Telegram gesendet.")

            bot_token, chat_id = _get_telegram_cfg()
            if bot_token and chat_id:
                try:
                    send_document(bot_token, chat_id, csv_path, caption)
                except Exception as e:
                    print(f"ⓘ Konnte CSV nicht senden: {e}")

        # final_sim für Chart + Excel bestimmen
        if is_auto and results and 'final_result' in results:
            final_sim_for_export = results['final_result']
            portfolio_files_for_export = results.get('optimal_portfolio', [])
        else:
            final_sim_for_export = final_report
            portfolio_files_for_export = []

        # Charts & Excel: Modus 3 automatisch, Modus 2 auf Anfrage
        do_export = is_auto
        if not is_auto:
            print()
            ans = input("  Charts & Excel erstellen und via Telegram senden? (j/n) [Standard: n]: ").strip().lower()
            do_export = ans in ('j', 'y', 'ja')

        if do_export:
            _generate_portfolio_chart(final_sim_for_export, portfolio_files_for_export,
                                      start_capital, start_date, end_date)
            _generate_trades_excel(final_sim_for_export, start_capital)

        print("=======================================================")
    else:
        if not is_auto:
            print("\nKeine Equity-Daten zum Exportieren vorhanden (Möglicherweise 0 Trades).")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--mode', default='1', type=str)
    args = parser.parse_args()
    print("\n--- Bitte Konfiguration für den Backtest festlegen ---")
    start_date = input(f"Startdatum (JJJJ-MM-TT) [Standard: 2023-01-01]: ") or "2023-01-01"
    end_date = input(f"Enddatum (JJJJ-MM-TT) [Standard: Heute]: ") or date.today().strftime("%Y-%m-%d")
    start_capital = int(input(f"Startkapital in USDT eingeben [Standard: 1000]: ") or 1000)

    max_dd_input = 100.0
    if args.mode == '3':
        max_dd_input = float(input(f"Gewünschten maximalen Drawdown in % eingeben [Standard: 30]: ") or 30.0)

    print("--------------------------------------------------")
    if args.mode == '2':
        run_shared_mode(is_auto=False, start_date=start_date, end_date=end_date, start_capital=start_capital, max_drawdown=100.0)
    elif args.mode == '3':
        run_shared_mode(is_auto=True, start_date=start_date, end_date=end_date, start_capital=start_capital, max_drawdown=max_dd_input)
    else:
        # Führt die Einzelanalyse (Modus 1) über den robusten Simulator-Pfad aus
        run_single_analysis_via_simulator(start_date=start_date, end_date=end_date, start_capital=start_capital)
